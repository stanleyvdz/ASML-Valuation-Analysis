"""
ASML Valuation Analysis — Chart Generator
Reads all data directly from ASML_DCF_Model.xlsx and produces
five publication-quality charts for the GitHub README.

Usage:
    cd notebooks
    python generate_charts.py

Output:
    ../outputs/charts/01_revenue_growth.png
    ../outputs/charts/02_margin_analysis.png
    ../outputs/charts/03_dcf_waterfall.png
    ../outputs/charts/04_peer_comparison.png
    ../outputs/charts/05_sensitivity_heatmap.png
"""

import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
import seaborn as sns
import openpyxl

# ─────────────────────────────────────────────
# 0.  CONFIG
# ─────────────────────────────────────────────

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "models", "ASML_DCF_Model.xlsx")
OUT_DIR    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "outputs", "charts")
os.makedirs(OUT_DIR, exist_ok=True)

# Colour palette — restrained, finance-grade
BLUE_DARK   = "#0A2540"
BLUE_MID    = "#1A56A4"
BLUE_LIGHT  = "#4A90D9"
GREY_DARK   = "#4A4A4A"
GREY_LIGHT  = "#E8EDF2"
GREEN       = "#1A7F5A"
RED         = "#C0392B"
AMBER       = "#D4860A"
WHITE       = "#FFFFFF"
ACCENT      = "#2EAADC"

def style():
    """Apply a clean, professional style to every chart."""
    plt.rcParams.update({
        "figure.facecolor"  : WHITE,
        "axes.facecolor"    : WHITE,
        "axes.edgecolor"    : GREY_LIGHT,
        "axes.spines.top"   : False,
        "axes.spines.right" : False,
        "axes.spines.left"  : True,
        "axes.spines.bottom": True,
        "axes.grid"         : True,
        "axes.grid.axis"    : "y",
        "grid.color"        : GREY_LIGHT,
        "grid.linewidth"    : 0.8,
        "font.family"       : "DejaVu Sans",
        "text.color"        : GREY_DARK,
        "axes.labelcolor"   : GREY_DARK,
        "xtick.color"       : GREY_DARK,
        "ytick.color"       : GREY_DARK,
        "xtick.labelsize"   : 10,
        "ytick.labelsize"   : 10,
        "axes.titlesize"    : 13,
        "axes.titleweight"  : "bold",
        "axes.titlecolor"   : BLUE_DARK,
        "axes.labelsize"    : 10,
    })

style()

def savefig(name):
    path = os.path.join(OUT_DIR, name)
    plt.savefig(path, dpi=180, bbox_inches="tight", facecolor=WHITE)
    plt.close()
    print(f"  ✓  Saved {name}")

# ─────────────────────────────────────────────
# 1.  READ DATA FROM EXCEL
# ─────────────────────────────────────────────

print("\nReading ASML_DCF_Model.xlsx …")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ── Historical Financials ──
ws_hist = wb["Historical Financials"]

YEARS_HIST = [2021, 2022, 2023, 2024, 2025]

# Row-by-row lookup helper
def find_row(ws, label, col=1, max_row=120):
    for r in range(1, max_row + 1):
        v = ws.cell(r, col).value
        if v and str(v).strip().lower() == label.strip().lower():
            return r
    return None

r_rev   = find_row(ws_hist, "Net Sales")
r_gp    = find_row(ws_hist, "Gross profit")
r_ebit  = find_row(ws_hist, "Operating income (EBIT / Income from operations)")
r_ni    = find_row(ws_hist, "Net income")
r_gm    = find_row(ws_hist, "Gross Margin %")
r_om    = find_row(ws_hist, "Operating Margin %")
r_fcf   = find_row(ws_hist, "Free cash flow (Operating CF – CapEx)")

# Values are in columns B-F (cols 2-6) for years 2021-2025
def read_hist_row(ws, row, ncols=5, start_col=2):
    return [ws.cell(row, start_col + i).value or 0 for i in range(ncols)]

revenue   = read_hist_row(ws_hist, r_rev)
gross_p   = read_hist_row(ws_hist, r_gp)
ebit      = read_hist_row(ws_hist, r_ebit)
net_inc   = read_hist_row(ws_hist, r_ni)
gm_pct    = [v * 100 for v in read_hist_row(ws_hist, r_gm)]
om_pct    = [v * 100 for v in read_hist_row(ws_hist, r_om)]
fcf_hist  = read_hist_row(ws_hist, r_fcf)

print(f"  Revenue 2021-2025: {revenue}")
print(f"  Gross margin:      {[round(v,1) for v in gm_pct]}")

# ── Projections ──
ws_proj = wb["Projections"]
YEARS_PROJ = [2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035]

r_pj_rev  = find_row(ws_proj, "Revenue")
r_pj_fcf  = find_row(ws_proj, "Unlevered Free Cash Flow (FCF)")

def read_proj_row(ws, row, ncols=10, start_col=2):
    return [ws.cell(row, start_col + i).value or 0 for i in range(ncols)]

rev_proj  = read_proj_row(ws_proj, r_pj_rev)
fcf_proj  = read_proj_row(ws_proj, r_pj_fcf)

# ── DCF Calculation ──
ws_dcf = wb["DCF Calculation"]

r_dcf_val   = find_row(ws_dcf, "DCF VALUE PER SHARE (EUR):")
r_curr_px   = find_row(ws_dcf, "Current Market Price (EUR):")
r_ev        = find_row(ws_dcf, "Enterprise Value:", max_row=40)
r_pv_fcfs   = find_row(ws_dcf, "Sum of PV of FCFs (2024-2033):")
r_pv_tv     = find_row(ws_dcf, "Plus: PV of Terminal Value:")
r_cash      = find_row(ws_dcf, "Plus: Cash & Cash Equivalents:")
r_debt      = find_row(ws_dcf, "Less: Total Debt:")
r_equity    = find_row(ws_dcf, "Equity Value:")

dcf_per_share = ws_dcf.cell(r_dcf_val, 4).value
curr_price    = ws_dcf.cell(r_curr_px, 4).value
pv_fcfs       = ws_dcf.cell(r_pv_fcfs, 4).value
pv_tv         = ws_dcf.cell(r_pv_tv,   4).value
cash_val      = ws_dcf.cell(r_cash,    4).value
debt_val      = abs(ws_dcf.cell(r_debt, 4).value)
equity_val    = ws_dcf.cell(r_equity,  4).value

print(f"  DCF per share: €{dcf_per_share:,.0f}")
print(f"  Current price: €{curr_price:,.0f}")

# ── Comparables ──
ws_comp = wb["Comparables"]

PEERS = []
for r in range(1, 50):
    company = ws_comp.cell(r, 1).value
    if company in ("ASML", "Applied Materials", "LAM Research", "KLA Corp"):
        ev_ebitda = ws_comp.cell(r, 3).value
        pe        = ws_comp.cell(r, 4).value
        gm        = ws_comp.cell(r, 6).value
        PEERS.append({
            "name": company,
            "ev_ebitda": ev_ebitda or 0,
            "pe": pe or 0,
            "gm": (gm or 0) * 100,
        })

print(f"  Peers found: {[p['name'] for p in PEERS]}")

# ── Sensitivity ──
ws_sens = wb["Sensitivity"]

# WACC values are in row 5, cols B-L (cols 2-12)
wacc_vals   = [ws_sens.cell(5, c).value for c in range(2, 13) if ws_sens.cell(5, c).value is not None]
# Growth values are in col A, rows 6-13
growth_vals = [ws_sens.cell(r, 1).value for r in range(6, 14) if ws_sens.cell(r, 1).value is not None]
# Data matrix: rows 6-13, cols 2-12
sens_matrix = []
for r in range(6, 6 + len(growth_vals)):
    row_data = [ws_sens.cell(r, c).value or 0 for c in range(2, 2 + len(wacc_vals))]
    sens_matrix.append(row_data)

sens_array = np.array(sens_matrix, dtype=float)
print(f"  Sensitivity matrix: {sens_array.shape[0]}×{sens_array.shape[1]}")

wb.close()
print("  Done reading workbook.\n")

# ─────────────────────────────────────────────
# CHART 1 — Revenue Growth (Historical + Projected)
# ─────────────────────────────────────────────

print("Building Chart 1: Revenue Growth …")

fig, ax = plt.subplots(figsize=(11, 5.5))

all_years = YEARS_HIST + YEARS_PROJ
all_rev   = revenue + rev_proj
x         = np.arange(len(all_years))

# Bars: historical vs projected
colors = [BLUE_DARK] * len(YEARS_HIST) + [BLUE_LIGHT] * len(YEARS_PROJ)
bars = ax.bar(x, all_rev, color=colors, width=0.65, zorder=3)

# Divider line between historical and projected
ax.axvline(len(YEARS_HIST) - 0.5, color=GREY_DARK, linewidth=1.2,
           linestyle="--", alpha=0.6)
ax.text(len(YEARS_HIST) - 0.42, max(all_rev) * 0.97,
        "Forecast →", fontsize=9, color=GREY_DARK, style="italic")

# Value labels on bars
for bar, val in zip(bars, all_rev):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(all_rev) * 0.012,
            f"€{val/1000:.0f}B" if val >= 1000 else f"€{val:.0f}M",
            ha="center", va="bottom", fontsize=7.5, color=BLUE_DARK, fontweight="bold")

ax.set_xticks(x)
ax.set_xticklabels([str(y) + ("E" if y >= 2026 else "") for y in all_years],
                   rotation=35, ha="right")
ax.set_ylabel("EUR (Millions)")
ax.set_title("ASML Net Sales — Historical & Projected (2021–2035)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"€{v:,.0f}"))

legend_handles = [
    mpatches.Patch(color=BLUE_DARK, label="Actual"),
    mpatches.Patch(color=BLUE_LIGHT, label="Projected"),
]
ax.legend(handles=legend_handles, frameon=False, fontsize=9)

# CAGR annotation
cagr_hist = (revenue[-1] / revenue[0]) ** (1 / (len(revenue) - 1)) - 1
cagr_proj = (rev_proj[-1] / rev_proj[0]) ** (1 / (len(rev_proj) - 1)) - 1
ax.text(0.01, 0.94,
        f"Historical CAGR ('21–'25): {cagr_hist:.1%}   |   Projected CAGR ('26–'35): {cagr_proj:.1%}",
        transform=ax.transAxes, fontsize=8.5, color=BLUE_DARK,
        bbox=dict(facecolor=GREY_LIGHT, edgecolor="none", pad=4, alpha=0.8))

plt.tight_layout()
savefig("01_revenue_growth.png")

# ─────────────────────────────────────────────
# CHART 2 — Margin Analysis (Historical)
# ─────────────────────────────────────────────

print("Building Chart 2: Margin Analysis …")

fig, ax = plt.subplots(figsize=(9, 5))

x = np.arange(len(YEARS_HIST))
w = 0.28

ax.bar(x - w, gm_pct,  width=w, color=BLUE_DARK,  label="Gross Margin",     zorder=3)
ax.bar(x,     om_pct,  width=w, color=BLUE_LIGHT,  label="Operating Margin", zorder=3)
ax.bar(x + w, [ni / re * 100 for ni, re in zip(net_inc, revenue)],
              width=w, color=ACCENT,    label="Net Margin",       zorder=3)

ax.set_xticks(x)
ax.set_xticklabels([str(y) for y in YEARS_HIST])
ax.set_ylabel("Margin (%)")
ax.set_title("ASML Profitability Margins (2021–2025)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
ax.legend(frameon=False, fontsize=9)

# Average labels
avg_gm = sum(gm_pct) / len(gm_pct)
avg_om = sum(om_pct) / len(om_pct)
ax.axhline(avg_gm, color=BLUE_DARK, linestyle=":", linewidth=1, alpha=0.5)
ax.axhline(avg_om, color=BLUE_LIGHT, linestyle=":", linewidth=1, alpha=0.5)
ax.text(len(YEARS_HIST) - 0.55, avg_gm + 0.6,
        f"Avg GM {avg_gm:.1f}%", fontsize=8, color=BLUE_DARK, alpha=0.8)
ax.text(len(YEARS_HIST) - 0.55, avg_om + 0.6,
        f"Avg OM {avg_om:.1f}%", fontsize=8, color=BLUE_LIGHT, alpha=0.8)

plt.tight_layout()
savefig("02_margin_analysis.png")

# ─────────────────────────────────────────────
# CHART 3 — DCF Valuation Waterfall
# ─────────────────────────────────────────────

print("Building Chart 3: DCF Waterfall …")

fig, ax = plt.subplots(figsize=(10, 5.5))

labels    = ["PV of FCFs\n(2026–2035)", "+ PV of\nTerminal Value",
            "= Enterprise\nValue", "+ Cash", "− Debt", "= Equity\nValue"]
bar_vals  = [pv_fcfs, pv_tv, pv_fcfs + pv_tv,
             cash_val, debt_val, equity_val]
bottoms   = [0, pv_fcfs, 0,
             pv_fcfs + pv_tv, pv_fcfs + pv_tv + cash_val - debt_val, 0]
colors_wf = [BLUE_LIGHT, BLUE_MID, BLUE_DARK, GREEN, RED, GREEN]

for i, (lbl, bv, bot, col) in enumerate(
        zip(labels, bar_vals, bottoms, colors_wf)):
    ax.bar(i, bv, bottom=bot, color=col, width=0.55, zorder=3,
           edgecolor=WHITE, linewidth=0.8)
    ax.text(i, bot + bv + max(bar_vals) * 0.015,
            f"€{bv/1000:.0f}B", ha="center", fontsize=8.5,
            fontweight="bold", color=BLUE_DARK)

# Per-share line
ax2 = ax.twinx()
ax2.set_ylim(0, equity_val / 1000 * 1.35)
ax2.set_yticks([])
ax.set_xticks(range(len(labels)))
ax.set_xticklabels(labels, fontsize=9.5)
ax.set_ylabel("EUR (Millions)")
ax.set_title("DCF Valuation Bridge — ASML")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"€{v:,.0f}"))

# Per-share annotation
ax.annotate(
    f"Implied value per share: €{dcf_per_share:,.0f}\n"
    f"vs. Current price: €{curr_price:,.0f}  "
    f"({'Overvalued' if dcf_per_share < curr_price else 'Undervalued'} "
    f"{abs(dcf_per_share/curr_price - 1):.0%})",
    xy=(5, equity_val), xytext=(3.6, equity_val * 1.05),
    fontsize=8.5, color=RED if dcf_per_share < curr_price else GREEN,
    fontweight="bold",
    arrowprops=dict(arrowstyle="->", color=GREY_DARK, lw=1.2),
    bbox=dict(facecolor=GREY_LIGHT, edgecolor="none", pad=4)
)

plt.tight_layout()
savefig("03_dcf_waterfall.png")

# ─────────────────────────────────────────────
# CHART 4 — Peer Comparison (EV/EBITDA & P/E)
# ─────────────────────────────────────────────

print("Building Chart 4: Peer Comparison …")

fig, axes = plt.subplots(1, 2, figsize=(12, 5.5))

peer_names  = [p["name"] for p in PEERS]
ev_ebitda_v = [p["ev_ebitda"] for p in PEERS]
pe_v        = [p["pe"] for p in PEERS]
gm_v        = [p["gm"] for p in PEERS]
bar_colors  = [BLUE_DARK if n == "ASML" else BLUE_LIGHT for n in peer_names]

for ax, metric, vals, title, ylabel in [
    (axes[0], "EV/EBITDA", ev_ebitda_v, "EV / EBITDA", "Multiple (×)"),
    (axes[1], "P/E Ratio",  pe_v,        "Price / Earnings", "Multiple (×)"),
]:
    bars = ax.bar(peer_names, vals, color=bar_colors, width=0.55, zorder=3,
                  edgecolor=WHITE, linewidth=0.8)
    for bar, val in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + max(vals) * 0.02,
                f"{val:.1f}×", ha="center", va="bottom",
                fontsize=9, fontweight="bold", color=BLUE_DARK)
    med = sorted(vals)[len(vals) // 2]
    ax.axhline(med, color=AMBER, linewidth=1.4, linestyle="--", alpha=0.8)
    ax.text(len(peer_names) - 0.45, med * 1.03,
            f"Median {med:.1f}×", fontsize=8, color=AMBER)
    ax.set_title(title)
    ax.set_ylabel(ylabel)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}×"))
    ax.set_xticklabels(peer_names, rotation=15, ha="right")

legend_handles = [
    mpatches.Patch(color=BLUE_DARK,  label="ASML"),
    mpatches.Patch(color=BLUE_LIGHT, label="Peers"),
    mpatches.Patch(color=AMBER,      label="Peer Median"),
]
axes[1].legend(handles=legend_handles, frameon=False, fontsize=9,
               loc="upper right")

fig.suptitle("ASML vs. Semiconductor Equipment Peers — Valuation Multiples",
             fontsize=13, fontweight="bold", color=BLUE_DARK, y=1.02)
plt.tight_layout()
savefig("04_peer_comparison.png")

# ─────────────────────────────────────────────
# CHART 5 — Sensitivity Heatmap
# ─────────────────────────────────────────────

print("Building Chart 5: Sensitivity Heatmap …")

fig, ax = plt.subplots(figsize=(13, 6))

wacc_labels   = [f"{v:.1%}" for v in wacc_vals]
growth_labels = [f"{v:.1%}" for v in growth_vals]

# Diverging palette centred on current market price
cmap = sns.diverging_palette(10, 130, s=80, l=45, as_cmap=True)
sns.heatmap(
    sens_array,
    ax=ax,
    annot=True, fmt=".0f",
    cmap=cmap,
    center=curr_price,
    xticklabels=wacc_labels,
    yticklabels=growth_labels,
    linewidths=0.4,
    linecolor=WHITE,
    cbar_kws={"label": "DCF Value Per Share (EUR)", "shrink": 0.85},
    annot_kws={"size": 8.5, "weight": "bold"},
)

ax.set_xlabel("WACC", fontsize=10, labelpad=8)
ax.set_ylabel("Terminal Growth Rate", fontsize=10, labelpad=8)
ax.set_title(
    f"Sensitivity Analysis — DCF Value Per Share (EUR)\n"
    f"Base case: WACC = 9.2%  |  Terminal Growth = 2.5%  |  "
    f"Current Price = €{curr_price:,.0f}",
    fontsize=11, fontweight="bold", color=BLUE_DARK, pad=12
)

# Highlight base case cell
try:
    base_row = [f"{v:.1%}" for v in growth_vals].index("2.5%")
    base_col = [f"{v:.1%}" for v in wacc_vals].index("9.2%")
    ax.add_patch(plt.Rectangle(
        (base_col, base_row), 1, 1,
        fill=False, edgecolor=BLUE_DARK, linewidth=2.5, zorder=5
    ))
except ValueError:
    pass

# Annotation: overvalued region
ax.text(0.01, -0.13,
        f"Red = Market price exceeds DCF value (stock overvalued) — "
        f"Base case implies −{abs(dcf_per_share/curr_price - 1):.0%} downside",
        transform=ax.transAxes, fontsize=8.5,
        color=RED, style="italic")

plt.tight_layout()
savefig("05_sensitivity_heatmap.png")

# ─────────────────────────────────────────────
# DONE
# ─────────────────────────────────────────────

print(f"\n{'='*55}")
print("All 5 charts written to outputs/charts/")
print(f"{'='*55}")
print("""
  01_revenue_growth.png     — Historical & projected revenue
  02_margin_analysis.png    — Gross / operating / net margins
  03_dcf_waterfall.png      — Valuation bridge to equity value
  04_peer_comparison.png    — EV/EBITDA and P/E vs peers
  05_sensitivity_heatmap.png — WACC × Terminal Growth sensitivity
""")
